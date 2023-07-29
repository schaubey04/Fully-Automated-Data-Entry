from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Data Entry")
root.resizable(False, False)
root.geometry("700x400+300+200")
root.configure(bg="#326273")

icon_image=PhotoImage(file="xlsx_logo.png")
root.iconphoto(False,icon_image)

file=pathlib.Path("Shivam_Backend_Data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Full Name"
    sheet["B1"]="PhoneNumber"
    sheet["C1"]="Age"
    sheet["D1"]="Gender"
    sheet["E1"]="Address"

    file.save("Shivam_Backend_Data.xlsx")

def submit():
    name=namevalue.get()
    contact=contactvalue.get()
    age=agevalue.get()
    gender=gender_combobox.get()
    address=AddressEntry.get(1.0,END)

    file=openpyxl.load_workbook("Shivam_Backend_Data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r"Shivam_Backend_Data.xlsx")



def clear():
    namevalue.set("")
    contactvalue.set("")
    agevalue.set("")
    AddressEntry.delete(1.0,END)

# icon


# heading

Label(root, text="Please Fill out this entry form:", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Label
Label(root, text="Name", font=23, bg="#326273", fg="white").place(x=50, y=100)
Label(root, text="Contact No: ", font=23, bg="#326273", fg="white").place(x=50, y=150)
Label(root, text="Age", font=23, bg="#326273", fg="white").place(x=50, y=200)
Label(root, text="Gender", font=23, bg="#326273", fg="white").place(x=370, y=200)
Label(root, text="Address", font=23, bg="#326273", fg="white").place(x=50, y=250)

# Entry
namevalue = StringVar()
contactvalue = StringVar()
agevalue = StringVar()

NameEntry = Entry(root, textvariable=namevalue, width=40, bd=2, font=20)
NameEntry.place(x=200, y=100)

contactEntry = Entry(root, textvariable=contactvalue, width=40, bd=2, font=20)
contactEntry.place(x=200, y=150)

ageEntry = Entry(root, textvariable=agevalue, width=13, bd=2, font=20)
ageEntry.place(x=200, y=200)

#Gender

gender_combobox=Combobox(root,values=["Male","Female","Transgender"],font="arial 10",width=18,state="r")
gender_combobox.place(x=450,y=205)
gender_combobox.set("Select Your Gender")

#Address

AddressEntry=Text(root,width=50,height=4,bd=4,wrap=WORD,font="arial 15")
AddressEntry.place(x=200,y=250)

#Button
Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)

root.mainloop()
